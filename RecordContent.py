from openpyxl import load_workbook

import ModelTest
import ModelInfoTest
import BaselineTest


def record_result():
    # 测试RNN模型
    month_list = [3, 6, 9, 12]
    emb_list = [500]
    hidden_list = [[300, 300], [300]]
    rnn_unit_list = ['gru']
    wb = load_workbook(filename='./test_result/32test_result.xlsx')
    sheet_names = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_names[0])
    ws_count = 26
    # for month in month_list:
    #     for emb in emb_list:
    #         for hidden in hidden_list:
    #             for rnn_unit in rnn_unit_list:
    #                 file_name = 'rnn' + str(len(hidden)) + '_' + str(emb) + 'emb_' + str(hidden) \
    #                             + 'hidden_' + '_' + rnn_unit + '_' + str(20) + 'epochs_' + str(
    #                     month) + 'month_info'
    #
    #                 sick_result = ModelInfoTest.test_model_probability(file_name + '.h5', month=month)
    #                 disease_result = ModelInfoTest.test_model_disease(file_name + '.h5', month=month)
    #                 ws.cell(column=1, row=ws_count, value=ws_count)
    #                 ws.cell(column=2, row=ws_count, value='RNN_INFO')
    #                 ws.cell(column=3, row=ws_count, value=emb)
    #                 ws.cell(column=4, row=ws_count, value=str(hidden))
    #                 ws.cell(column=5, row=ws_count, value=rnn_unit)
    #                 ws.cell(column=6, row=ws_count, value=month)
    #                 ws.cell(column=7, row=ws_count, value=sick_result[0])
    #                 ws.cell(column=8, row=ws_count, value=sick_result[1])
    #                 ws.cell(column=9, row=ws_count, value=sick_result[2])
    #                 ws.cell(column=10, row=ws_count, value=sick_result[3])
    #                 ws.cell(column=11, row=ws_count, value=sick_result[4])
    #                 ws.cell(column=12, row=ws_count, value=disease_result[0])
    #                 ws.cell(column=13, row=ws_count, value=disease_result[1])
    #                 ws.cell(column=14, row=ws_count, value=disease_result[2])
    #                 ws_count += 1
    #                 wb.save('./test_result/32test_result.xlsx')
    for month in month_list:
        for emb in emb_list:
            for hidden in hidden_list:
                for rnn_unit in rnn_unit_list:
                    file_name = 'rnn' + str(len(hidden)) + '_' + str(emb) + 'emb_' + str(hidden) \
                                + 'hidden_' + '_' + rnn_unit + '_' + str(20) + 'epochs_' + str(
                        month) + 'month_info'

                    sick_result = ModelInfoTest.test_model_probability(file_name + '.h5', month=month)
                    disease_result = ModelInfoTest.test_model_disease(file_name + '.h5', month=month)
                    ws.cell(column=1, row=ws_count, value=ws_count)
                    ws.cell(column=2, row=ws_count, value='RNN_INFO')
                    ws.cell(column=3, row=ws_count, value=emb)
                    ws.cell(column=4, row=ws_count, value=str(hidden))
                    ws.cell(column=5, row=ws_count, value=rnn_unit)
                    ws.cell(column=6, row=ws_count, value=month)
                    ws.cell(column=7, row=ws_count, value=sick_result[0])
                    ws.cell(column=8, row=ws_count, value=sick_result[1])
                    ws.cell(column=9, row=ws_count, value=sick_result[2])
                    ws.cell(column=10, row=ws_count, value=sick_result[3])
                    ws.cell(column=11, row=ws_count, value=sick_result[4])
                    ws.cell(column=12, row=ws_count, value=disease_result[0])
                    ws.cell(column=13, row=ws_count, value=disease_result[1])
                    ws.cell(column=14, row=ws_count, value=disease_result[2])
                    ws_count += 1
                    wb.save('./test_result/32test_result.xlsx')
    # for month in month_list:
    #     for emb in emb_list:
    #         for hidden in hidden_list:
    #             for rnn_unit in rnn_unit_list:
    #                 file_name = 'rnn' + str(len(hidden)) + '_' + str(emb) + 'emb_' + str(hidden) \
    #                             + 'hidden_' + '_' + rnn_unit + '_' + str(20) + 'epochs_' + str(
    #                     month) + 'month'
    #
    #                 sick_result = ModelTest.test_model_probability(file_name + '.h5', month=month)
    #                 disease_result = ModelTest.test_model_disease(file_name + '.h5', month=month)
    #                 ws.cell(column=1, row=ws_count, value=ws_count)
    #                 ws.cell(column=2, row=ws_count, value='RNN')
    #                 ws.cell(column=3, row=ws_count, value=emb)
    #                 ws.cell(column=4, row=ws_count, value=str(hidden))
    #                 ws.cell(column=5, row=ws_count, value=rnn_unit)
    #                 ws.cell(column=6, row=ws_count, value=month)
    #                 ws.cell(column=7, row=ws_count, value=sick_result[0])
    #                 ws.cell(column=8, row=ws_count, value=sick_result[1])
    #                 ws.cell(column=9, row=ws_count, value=sick_result[2])
    #                 ws.cell(column=10, row=ws_count, value=sick_result[3])
    #                 ws.cell(column=11, row=ws_count, value=sick_result[4])
    #                 ws.cell(column=12, row=ws_count, value=disease_result[0])
    #                 ws.cell(column=13, row=ws_count, value=disease_result[1])
    #                 ws.cell(column=14, row=ws_count, value=disease_result[2])
    #                 ws_count += 1
    #                 wb.save('./test_result/32test_result.xlsx')
    # for month in month_list:
    #     sick_result = BaselineTest.test_probability('lr_10epochs_' + str(month) + 'month.h5', month=month)
    #     disease_result = BaselineTest.test_disease('lr_10epochs_' + str(month) + 'month.h5', month=month)
    #     ws.cell(column=1, row=ws_count, value=ws_count)
    #     ws.cell(column=2, row=ws_count, value='LR')
    #     ws.cell(column=6, row=ws_count, value=month)
    #     ws.cell(column=7, row=ws_count, value=sick_result[0])
    #     ws.cell(column=8, row=ws_count, value=sick_result[1])
    #     ws.cell(column=9, row=ws_count, value=sick_result[2])
    #     ws.cell(column=10, row=ws_count, value=sick_result[3])
    #     ws.cell(column=11, row=ws_count, value=sick_result[4])
    #     ws.cell(column=12, row=ws_count, value=disease_result[0])
    #     ws.cell(column=13, row=ws_count, value=disease_result[1])
    #     ws.cell(column=14, row=ws_count, value=disease_result[2])
    #     ws_count += 1
    #     wb.save('./test_result/32test_result.xlsx')
    #     sick_result = BaselineTest.test_probability('mlp_10epochs_' + str(month) + 'month.h5', month=month)
    #     disease_result = BaselineTest.test_disease('mlp_10epochs_' + str(month) + 'month.h5', month=month)
    #     ws.cell(column=1, row=ws_count, value=ws_count)
    #     ws.cell(column=2, row=ws_count, value='MLP')
    #     ws.cell(column=6, row=ws_count, value=month)
    #     ws.cell(column=7, row=ws_count, value=sick_result[0])
    #     ws.cell(column=8, row=ws_count, value=sick_result[1])
    #     ws.cell(column=9, row=ws_count, value=sick_result[2])
    #     ws.cell(column=10, row=ws_count, value=sick_result[3])
    #     ws.cell(column=11, row=ws_count, value=sick_result[4])
    #     ws.cell(column=12, row=ws_count, value=disease_result[0])
    #     ws.cell(column=13, row=ws_count, value=disease_result[1])
    #     ws.cell(column=14, row=ws_count, value=disease_result[2])
    #     ws_count += 1
    #     wb.save('./test_result/32test_result.xlsx')


if __name__ == "__main__":
    record_result()
