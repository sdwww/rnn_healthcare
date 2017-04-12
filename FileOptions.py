import pickle
import datetime
import re

import xlrd, xlwt
import numpy as np
import DBOptions
from openpyxl.reader.excel import load_workbook


# 从list中选择最长的item
def get_longseq(seqs):
    longseq = ''
    for i in seqs:
        if len(i) > len(longseq):
            longseq = i
    return longseq


# 将string类型转换为date类型
def str2date(str_item):
    date_split = str_item.split('-')
    date_item = datetime.date(int(date_split[0]), int(date_split[1]), int(date_split[2]))
    return date_item


# 将数据存储成pkl类型
def dump_pkl(data, filename):
    output = open("./data_pkl/" + filename, 'wb')
    pickle.dump(data, output, -1)
    output.close()


# 读取pkl类型的数据
def load_pkl(filename):
    f = open("./data_pkl/" + filename, 'rb')
    data = pickle.load(f)
    return data


def normal_leven(str1, str2):
    len_str1 = len(str1) + 1
    len_str2 = len(str2) + 1
    # create matrix
    matrix = [0 for n in range(len_str1 * len_str2)]
    # init x axis
    for i in range(len_str1):
        matrix[i] = i
    # init y axis
    for j in range(0, len(matrix), len_str1):
        if j % len_str1 == 0:
            matrix[j] = j // len_str1

    for i in range(1, len_str1):
        for j in range(1, len_str2):
            if str1[i - 1] == str2[j - 1]:
                cost = 0
            else:
                cost = 1
            matrix[j * len_str1 + i] = min(matrix[(j - 1) * len_str1 + i] + 1,
                                           matrix[j * len_str1 + (i - 1)] + 1,
                                           matrix[(j - 1) * len_str1 + (i - 1)] + cost)

    return matrix[-1]


# 该方法生成icd10与疾病的键值对列表
def get_ICD10_jbbm():
    book = xlrd.open_workbook(r'./data_xls/2013版疾病ICD10码.xls', formatting_info=True)
    sheet0 = book.sheet_by_index(0)
    nrows = sheet0.nrows
    ICD10_jbbm = {}
    for i in range(1, nrows):
        rows = sheet0.row_values(i)
        ICD10_jbbm[rows[1]] = rows[0]
    return ICD10_jbbm


# 获取药典中药品和药品分类的键值对
def get_drug_categ():
    book = xlrd.open_workbook(r'./data_xls/医典合集.xls', formatting_info=True)
    sheet0 = book.sheet_by_index(0)
    nrows = sheet0.nrows
    cate = ''
    drug_categ = {}
    for i in range(nrows):
        rows = sheet0.row_values(i)
        xfx = sheet0.cell_xf_index(i, 0)
        xf = book.xf_list[xfx]
        bgx = xf.background.pattern_colour_index
        if bgx != 64:
            cate = rows[0]
        else:
            drug_categ[rows[0]] = cate
    return drug_categ


# 获取数据库中所有的商品名
def get_all_spm():
    book = xlrd.open_workbook(filename='./data_xlsx/SPM_COUNT.xlsx')
    sheet0 = book.sheet_by_index(0)
    nrows = sheet0.nrows
    all_spm_list = []
    for i in range(1, nrows):
        rows = sheet0.row_values(i)
        all_spm_list.append(rows[0])
    return all_spm_list


# 获取数据库中所有的疾病名称
def get_all_jbmc():
    book = xlrd.open_workbook(r'./data_xls/JBMC_COUNT.xls', formatting_info=True)
    sheet0 = book.sheet_by_index(0)
    nrows = sheet0.nrows
    all_jbmc_list = []
    for i in range(1, nrows):
        rows = sheet0.row_values(i)
        all_jbmc_list.append(rows[0])
    return all_jbmc_list


# 获取数据库中所有的疾病名称和疾病类别键值对
def get_jbbm_jblb():
    book = xlrd.open_workbook(r'./data_xls/JBBM_JBLB.xls', formatting_info=True)
    sheet0 = book.sheet_by_index(0)
    nrows = sheet0.nrows
    all_jbbm_jblb = {}
    for i in range(1, nrows):
        rows = sheet0.row_values(i)
        all_jbbm_jblb[rows[0]] = rows[2]
    return all_jbbm_jblb


def get_spm_spmlb():
    drug_categ = get_drug_categ()
    spm_drug_dict = load_pkl('spm_drug_dict')
    drug_list = []
    for i in spm_drug_dict:
        if spm_drug_dict[i] not in drug_list:
            drug_list.append(spm_drug_dict[i])
    spm_spmlb = {}
    for i in drug_list:
        spm_spmlb[i] = drug_categ[i]
    spmlb_list = []
    for i in spm_spmlb:
        if spm_spmlb[i] not in spmlb_list:
            spmlb_list.append(spm_spmlb[i])
    print(len(spmlb_list))
    count = 0
    spmlb_index = {}
    for i in spmlb_list:
        spmlb_index[i] = count
        count += 1
    spm_spmlbindex = {}
    for i in spm_spmlb:
        spm_spmlbindex[i] = spmlb_index[spm_spmlb[i]]
    return spm_spmlbindex


# 获取数据库商品名与药典中的药名的对应关系
def get_spm_drug():
    drug_categ = get_drug_categ()
    all_spm_list = get_all_spm()
    count = 0
    spm_drug = {}
    for i in all_spm_list:
        for j in list(drug_categ.keys()):
            if i.find(j) != -1:
                if i in spm_drug:
                    spm_drug[i].append(j)
                else:
                    spm_drug[i] = [j]
                count += 1
    spm_drug_new = {}
    for i in spm_drug:
        if len(spm_drug[i]) >= 2:
            spm_drug_new[i] = get_longseq(spm_drug[i])
        else:
            spm_drug_new[i] = spm_drug[i][0]
    spm_drug = spm_drug_new
    dump_pkl(spm_drug, 'spm_drug_dict')


# 获取数据库的信息准确的疾病名称和icd10键值对
def get_jbmc_icd10():
    jbmc_icd10 = {}
    all_jbmc_list = get_all_jbmc()

    icd10_jbmc = get_ICD10_jbbm()
    for i in all_jbmc_list:
        for j in icd10_jbmc:
            if i == j:
                jbmc_icd10[i] = icd10_jbmc[i]
    dump_pkl(jbmc_icd10, 'jbmc_icd10_dict')


# 创建pkl文件，内容为包含相关icd10和药品的部分数据库记录
def create_allinfo_pkl(cursor):
    spm_drug = load_pkl('spm_drug_dict')

    spm_list = []
    for i in spm_drug:
        spm_list.append(i)

    jbmc_icd10 = load_pkl('jbmc_icd10_dict')
    jbmc_list = list(jbmc_icd10.keys())

    all_info = DBOptions.getAllInfo(cursor, spm_list, jbmc_list)
    dump_pkl(all_info, 'all_info')


# 创建数据集
def createDataset(cursor):
    sql = 'select grbh_index,xh_index,period,jbbm_index,spm_index from DATA_ANALYSIS_FINAL where xh_INDEX!=0'
    all_info = DBOptions.getSQL(sql, cursor=cursor)
    dataset_x = np.zeros((16700, 96, 2094), dtype='int32')
    print('CODE数据集大小：', dataset_x.shape)
    for i in all_info:
        dataset_x[i[0], -i[1], i[3]] = 1
        dataset_x[i[0], -i[1], i[4]] = 1

    dataset_t = np.zeros((16700, 96, 1), dtype='int16')
    print('TIME数据集大小：', dataset_t.shape)
    for i in all_info:
        dataset_t[i[0], -i[1], 0] = i[2]

    sql = 'select grbh_index,xh_index,jblb_index,spmlb_index from DATA_ANALYSIS_FINAL where xh_INDEX=0'
    all_info = DBOptions.getSQL(sql, cursor=cursor)
    label_y = np.zeros((16700, 301), dtype='int16')
    print('CODE标签大小：', label_y.shape)
    for i in all_info:
        label_y[i[0], i[2]] = 1
        label_y[i[0], i[3]] = 1
    label_t = np.zeros((16700, 1), dtype='int16')
    print('TIME标签大小：', label_t.shape)
    for i in all_info:
        label_t[i[0], 0] = i[2]
    dataset_x_train = dataset_x[:14000]
    dataset_x_test = dataset_x[14000:]
    dataset_t_train = dataset_t[:14000]
    dataset_t_test = dataset_t[14000:]
    label_y_train = label_y[:14000]
    label_y_test = label_y[14000:]
    label_t_train = label_t[:14000]
    label_t_test = label_t[14000:]
    np.savez_compressed('./data_npz/dataset_x_train.npz', dataset_x_train)
    np.savez_compressed('./data_npz/dataset_x_test.npz', dataset_x_test)
    np.savez_compressed('./data_npz/dataset_t_train.npz', dataset_t_train)
    np.savez_compressed('./data_npz/dataset_t_test.npz', dataset_t_test)
    np.savez_compressed('./data_npz/label_y_train.npz', label_y_train)
    np.savez_compressed('./data_npz/label_y_test.npz', label_y_test)
    np.savez_compressed('./data_npz/label_t_train.npz', label_t_train)
    np.savez_compressed('./data_npz/label_t_test.npz', label_t_test)


# 原计划制定规则删除数据
def revise_DELETE():
    wb = load_workbook(filename='./data_xlsx/SPM_COUNT.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    delete_count = 0
    for rx in range(2, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w2 = int(ws.cell(row=rx, column=2).value)
        if not w1 or w1.find('一次性') != -1 or w1.find('注射器') != -1 \
                or w1.find('酒精') != -1 or w1.find('留置针') != -1 or w1.find('病号') == 0 \
                or w1.find('病人') == 0 or w1.find('床单') != -1 or w1.find('护理') != -1 \
                or w1.find('人间') != -1 or w1.find('静脉输液') != -1 or w1.find('静脉注射') != -1 \
                or w1.find('基础收费') != -1 or w1.find('疾病健康教育') != -1 or w1.find('建床费') != -1 \
                or w1.find('巡诊') != -1 or w1.find('家庭治疗') != -1 or w1.find('煎药') == 0 \
                or w1.find('健康咨询') != -1 or w1.find('健康档案') != -1 or w1.find('诊查费') != -1 \
                or w1 == '其他' or w1 == '其它' or w1.find('手术材料') == 0 \
                or w1.find('手术包') == 0 or w1.find('手术刀') == 0 or w1.find('手术垫') == 0 \
                or w1.find('手术费') == 0 or w1.find('手术巾') == 0 or w1 == '刷套' \
                or w1.find('特定计算机') == 0 or w1 == '体检费' or w1.find('体温') == 0 \
                or w1.find('图文') == 0 or w1.find('一室两床') == 0 or w1.find('诊疗费') != -1 \
                or w1.find('一室两床') == 0 or w1.find('一室三床') == 0 or w1 == '主任医师' \
                or w1.find('统筹') == 0 or w1.find('取暖') != -1 or w1.find('空针') != -1 or w1.find('床位') != -1:
            ws.cell(column=3, row=rx, value=1)
            delete_count += w2
    wb.save('./data_xlsx/SPM_COUNT.xlsx')
    print('无用项目数：', delete_count)


# 维护SPM_COUNT.xlsx，添加药典中的药品与原来的商品名匹配
def match_drug():
    spm_drug_dict = load_pkl('spm_drug_dict')
    wb = load_workbook(filename='./data_xlsx/SPM_COUNT.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    delete_count = 0
    for rx in range(2, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w2 = int(ws.cell(row=rx, column=2).value)
        for i in spm_drug_dict:
            if w1 and w1 == i:
                ws.cell(column=4, row=rx, value=spm_drug_dict[i])
                delete_count += w2
    wb.save('./data_xlsx/SPM_COUNT.xlsx')
    print('药典项目数：', delete_count)


# 维护SPM_COUNT.xlsx，添加药典中的药品的药品分类
def match_drug_categ():
    drug_categ_dict = get_drug_categ()
    wb = load_workbook(filename='./data_xlsx/SPM_COUNT.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    delete_count = 0
    for rx in range(2, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w2 = int(ws.cell(row=rx, column=2).value)
        w4 = ws.cell(row=rx, column=4).value
        if w4:
            ws.cell(column=5, row=rx, value=drug_categ_dict[w4])
            delete_count += w2
    wb.save('./data_xlsx/SPM_COUNT.xlsx')
    print('药典项目数：', delete_count)


# 对JBBM_JBMC_COUNT.xlsx进行修改，将与icd10表格完全匹配的疾病名称写入到xlsx文件中
def match_jbbm_jbmc():
    all_jbbm_icd10_dict = get_ICD10_jbbm()
    wb = load_workbook(filename='./data_xlsx/JBBM_JBMC_COUNT.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    delete_count = 0
    for rx in range(2, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w2 = ws.cell(row=rx, column=2).value
        w3 = ws.cell(row=rx, column=3).value
        for i in all_jbbm_icd10_dict:
            if w2 == i:
                ws.cell(column=4, row=rx, value=all_jbbm_icd10_dict[i])
                ws.cell(column=5, row=rx, value=i)
                delete_count += int(w3)
    print(delete_count)
    wb.save('./data_xlsx/JBBM_JBMC_COUNT.xlsx')
    print('icd10项目数：', delete_count)


# 对于记录数小于1000的jbbm和jbmc匹配，保持疾病编码的前五位不动，使用最小编辑距离自动进行填充
def auto_match_jbbm_jbmc():
    all_jbbm_icd10_dict = get_ICD10_jbbm()
    wb = load_workbook(filename='./data_xlsx/JBBM_JBMC_COUNT.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    count = 0
    pattern = re.compile(r'^[A-Z][0-9]{2}[.][0-9]{3}')
    for rx in range(2, ws.max_row + 1):
        w1 = str(ws.cell(row=rx, column=1).value)
        w2 = ws.cell(row=rx, column=2).value
        w3 = ws.cell(row=rx, column=3).value
        w4 = ws.cell(row=rx, column=4).value
        m = pattern.match(w1)
        if not w4 and m:
            match_dict = {}
            for i in all_jbbm_icd10_dict:
                if all_jbbm_icd10_dict[i][0:5] == w1[0:5]:
                    match_dict[i] = all_jbbm_icd10_dict[i]
            short_len = 100
            w4 = ''
            w5 = ''
            for i in match_dict:
                if normal_leven(i, w2) < short_len:
                    w4 = match_dict[i]
                    w5 = i
                    short_len = normal_leven(i, w2)
            if short_len <= 2:
                ws.cell(column=4, row=rx, value=w4)
                ws.cell(column=5, row=rx, value=w5)
                count += 1
    wb.save('./data_xlsx/JBBM_JBMC_COUNT.xlsx')
    print(count)


# 对JBBM_JBMC_COUNT.xlsx进行修改，将与icd10表格完全匹配的疾病名称写入到xlsx文件中
def show_info_ibbm():
    wb = load_workbook(filename='./data_xlsx/JBBM_JBMC_COUNT.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    delete_count = 0
    drug_list = []
    for rx in range(2, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w4 = ws.cell(row=rx, column=4).value
        if w4 and w4 not in drug_list:
            drug_list.append(w4)
    wb = load_workbook(filename='./data_xlsx/章节名称及代码（ICD-10）.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    for rx in range(1, ws.max_row + 1):
        count = 0
        w1 = ws.cell(row=rx, column=1).value
        w2 = ws.cell(row=rx, column=2).value
        start_int = int(w1[1:3])
        end_int = int(w1[5:7])
        for i in drug_list:
            if w1.find(i[0]) != -1 and int(i[1:3]) >= start_int and int(i[1:3]) <= end_int:
                print(i, w1, w2)
                delete_count += 1
                count += 1
    print(delete_count)
    print(len(drug_list))
    # wb.save('./data_xlsx/JBBM_JBMC_COUNT.xlsx')


# 对JBBM_JBMC_COUNT.xlsx进行修改，添加疾病类别
def match_jbbmlb():
    wb = load_workbook(filename='./data_xlsx/JBBM_JBMC_COUNT.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    delete_count = 0
    drug_list = []
    for rx in range(2, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w4 = ws.cell(row=rx, column=4).value
        if w4 and w4 not in drug_list:
            drug_list.append(w4)
    drug_categ_dict = {}
    wb = load_workbook(filename='./data_xlsx/3位代码类目表（ICD-10）.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    for rx in range(2, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w2 = ws.cell(row=rx, column=2).value
        for i in drug_list:
            if w1[0:3] == i[0:3]:
                drug_categ_dict[i] = w2
    wb = load_workbook(filename='./data_xlsx/JBBM_JBMC_COUNT.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    for rx in range(2, ws.max_row + 1):
        w4 = ws.cell(row=rx, column=4).value
        if w4:
            ws.cell(column=6, row=rx, value=drug_categ_dict[w4])
            print(w4)
    wb.save('./data_xlsx/JBBM_JBMC_COUNT.xlsx')


# 对JBBM_JBMC_COUNT.xlsx进行测试使用
def jbbm_jbmc_count_test():
    wb = load_workbook(filename='./data_xlsx/JBBM_JBMC_COUNT.xlsx')
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    drug_list = []
    for rx in range(2, ws.max_row + 1):
        w1 = ws.cell(row=rx, column=1).value
        w4 = ws.cell(row=rx, column=4).value
        if w4 and w4 not in drug_list:
            drug_list.append(w4)
    print(len(drug_list))
